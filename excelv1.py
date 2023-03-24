from pathlib import Path
from openpyxl import load_workbook, workbook
import datetime
import pandas as pd

# current_date
todays_date = str(datetime.datetime.now().strftime("%m.%d.%Y"))
todays_date1 = str(datetime.datetime.now().strftime("%m/%d/%Y"))


wb = load_workbook(r'/Users/joelbayingana/Downloads/Daily Overview(FKA 2hr).xlsx')

ws = wb.active

target = wb.copy_worksheet(ws)

target.title = todays_date


ws['c2'] = todays_date1






excel_file_1= ('/Users/joelbayingana/Downloads/TFSW Master In Process (1).xlsx')
excel_file_2 = ('/Users/joelbayingana/Downloads/Daily Overview(FKA 2hr).xlsx')


df_TwoHRsummary= pd.read_excel(excel_file_1,sheet_name='2HR Summary')

#d2=excel_file_1.merge(excel_file_2,on='QT', how='left')



   # Order_Totals = ws.cell(row=i, column=2).value

#df_2hr = pd.read_excel(excel_file_2,sheet_name= todays_date)

#sht= wb.sheets(todays_date)

#wb.save(r'/Users/joelbayingana/Downloads/Daily Overview(FKA 2hr).xlsx')

#print(d2)