import pandas as pd

import datetime as dt

from openpyxl import workbook, load_workbook

wb=load_workbook('daily_overview(FKA 2hr).xlsx')

wb.create_sheet("Test")
print(wb.sheetnames)

wb.save('daily_overview(FKA 2hr).xlsx')
#ws = wb.active
#print(ws['C2'].value)
