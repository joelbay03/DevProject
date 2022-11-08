
from openpyxl import workbook, load_workbook
import os
import datetime
import requests

#current_date = dt.date.today()
todays_date = str(datetime.datetime.now().strftime("%m.%d.%y"))

wb=load_workbook('daily_overview(FKA 2hr).xlsx')

ws=wb.active

target = wb.copy_worksheet(ws)

target.title= todays_date

#ws= todays_date

print(wb.sheetnames)

wb.save('daily_overview(FKA 2hr).xlsx')

