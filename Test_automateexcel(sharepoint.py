from openpyxl import workbook, load_workbook
import os
import datetime
import requests
import base64


#current_date = dt.date.today()

todays_date = str(datetime.datetime.now().strftime("%m.%d.%Y"))

wb=load_workbook(r'C:\Users\JBayingana\Taylor Farms\Brent Jones - Planning - In Process\Daily Overview(FKA 2hr).xlsx')

ws=wb.active

target = wb.copy_worksheet(ws)

target.title= todays_date



#ws= todays_date

print(wb.sheetnames)

wb.save(r'C:\Users\JBayingana\Taylor Farms\Brent Jones - Planning - In Process\Daily Overview(FKA 2hr).xlsx')


